import os
import openpyxl
import csv


ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
data_params_xml = os.path.join(ROOT_DIR, "data_params.xlsx")
data_params_csv = os.path.join(ROOT_DIR, "data_params_csv.csv")

def getXmlData():
    list = []
    wb = openpyxl.load_workbook(data_params_xml)
    print(wb.sheetnames) # Какие есть sheets

    sheet = wb['Profile_data'] # Выбираем с какие sheet работать

    rows = sheet.max_row # Считываем максимальное количество строк
    cols = sheet.max_column # Считываем маскимальное количество колонок

    big_list = []
    for r in range(2, rows+1): # Чтобы не считывать название колонок задаем шаг 2
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value

        tuple = (username, password)
        list.append(tuple)

    print(list)
    return list

def getCSVData():
    rows = []
    with open((os.path.join(data_params_csv)), "r") as csv_file:
        reader = csv.reader(csv_file) # Get csv file
        next(reader) #Skip headers, to read data
        for row in reader:
            rows.append(row)
        return rows
